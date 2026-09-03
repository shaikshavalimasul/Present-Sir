import io
import openpyxl
import xlrd
from flask import Blueprint, request, jsonify
from ..db import get_db

roster_bp = Blueprint("roster", __name__)

def _parse_excel(file_bytes, filename):
    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [{headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)} for row in ws.iter_rows(min_row=2, values_only=True)]
    else:
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        headers = [str(ws.cell_value(0, c)).strip().lower() for c in range(ws.ncols)]
        rows = [{headers[c]: str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)} for r in range(1, ws.nrows)]
    return headers, rows

@roster_bp.post("/upload")
def upload_roster():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    filename = f.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        return jsonify({"error": "Only .xlsx or .xls supported"}), 400
    file_bytes = f.read()
    try:
        headers, rows = _parse_excel(file_bytes, filename)
    except Exception as e:
        return jsonify({"error": f"Failed to parse file: {str(e)}"}), 400

    if "name" not in headers or "admission_number" not in headers:
        return jsonify({"error": "Excel must have 'name' and 'admission_number' columns"}), 400

    db = get_db()
    existing = {s["admission_number"] for s in db.table("students").select("admission_number").execute().data}

    to_insert = []
    skipped = 0
    for row in rows:
        adm = row.get("admission_number", "").strip()
        name = row.get("name", "").strip()
        if not adm or not name:
            continue
        if adm in existing:
            skipped += 1
            continue
        to_insert.append({"name": name, "admission_number": adm})
        existing.add(adm)

    if to_insert:
        db.table("students").insert(to_insert).execute()

    return jsonify({"inserted": len(to_insert), "skipped": skipped})

@roster_bp.get("/students")
def get_all_students():
    db = get_db()
    res = db.table("students").select("id, name, admission_number, has_registered").order("name").execute()
    return jsonify(res.data)

@roster_bp.post("/students/<student_id>/reset")
def reset_student(student_id):
    db = get_db()
    res = db.table("students").select("id").eq("id", student_id).execute()
    if not res.data:
        return jsonify({"error": "Student not found"}), 404
    db.table("students").update({"has_registered": False, "fingerprint_hash": None}).eq("id", student_id).execute()
    return jsonify({"message": "Student registration reset. They can now re-register."})
